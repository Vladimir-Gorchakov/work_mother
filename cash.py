import openpyxl
from openpyxl.styles import Border, Side, Alignment

import argparse
import datetime
from glob import glob
import os
import logging

number_format = '_-* #,##0.00\\ _₽_-;\\-* #,##0.00\\ _₽_-;_-* "-"??\\ _₽_-;_-@_-'
NUM_TAVRS = 3


def make_shablon(start, sheet, info, empty = False, color = None):

    if empty: 
        total_len = 14
    else:
        total_len = len(info['income']) + len(info['outcome']) + 11

    end = start + total_len
    date = info['date']

    sheet[f"B{start}"].value = 'Остаток на начало дня'
    sheet.merge_cells(f'B{start}:B{start+1}')

    # перенос суммы с прошлого дня
    sheet[f"C{start}"].value = f'=C{start-2}'
    sheet[f"C{start}"].number_format = number_format

    sheet[f"B{end - 3}"].value = 'сумма инкасации'
    sheet[f"C{end - 3}"].value = f'=B{end - 1}-C{end - 1}'
    sheet[f"C{end - 3}"].number_format = number_format

    sheet[f"B{end-2}"].value = 'приход'
    sheet[f"C{end-2}"].value = 'расход'
    sheet[f"B{end-1}"].value = f'=_xlfn.SUM(D{start}:D{end})'
    sheet[f"C{end-1}"].value = f'=_xlfn.SUM(E{start}:E{end})'
    sheet[f"B{end-1}"].number_format = number_format
    sheet[f"C{end-1}"].number_format = number_format

    sheet[f"B{end}"].value = "Остаток на конец дня"
    sheet.merge_cells(f'B{end}:B{end+1}')
    sheet[f"C{end}"].value = f'=C{start}+B{end-1}-C{end-1}'
    sheet[f"C{end}"].number_format = number_format


    if empty:
        for i in range(NUM_TAVRS):
            sheet[f"F{start+i}"].value = f'Выручка КАФЕ Т-{i+1}'
            sheet[f"G{start+i}"].value = 'кафе'
            sheet[f"H{start+i}"].value = 'выпечка'
            sheet[f"I{start+i}"].value = f'т{i+1}'

    else:
        for i, val in enumerate(info['names']):
            sheet[f"F{start+i}"].value = val
            sheet[f"H{start+i}"].value = 'выручка'

    # Добавляем дату
    for num in range(start, end + 2):
        sheet[f"A{num}"].value = date
        sheet[f"A{num}"].number_format = 'dd/mm/yy;@'


    # Красим, добавляем границы жирные края и тд

    if empty:
        color = openpyxl.styles.PatternFill(start_color=color,
                   fill_type='solid')
        
    
    matrix_cells = sheet[f'A{start}:K{end+1}']
    len_matrix = len(matrix_cells)
    len_cells = len(matrix_cells[0])

    for i, cells in enumerate(matrix_cells):
        for j, cell in enumerate(cells):

            if empty: 
                cell.fill = color

            cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin' if len_cells - 1 > j else 'thick'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin' if len_matrix - 1 > i else 'thick')
                    )


    return end+2


def insert_numbers(start_row, sheet_main, info):

    for i, val in enumerate(info['revenue']):
        sheet_main[f"D{start_row + i}"] = val[0]
        sheet_main[f"D{start_row + i}"].number_format = number_format
        sheet_main[f"G{start_row + i}"] = val[1]
        sheet_main[f"I{start_row + i}"] = info['tavr']

    for i, val in enumerate(info['acquiring']):
        sheet_main[f"J{start_row + i}"] = val
        sheet_main[f"J{start_row + i}"].number_format = number_format
        sheet_main[f"K{start_row + i}"] = f'=D{start_row + i}+J{start_row + i}'
        sheet_main[f"K{start_row + i}"].number_format = number_format

    start_row += len(info['names'])
    for i, val in enumerate(info['outcome']):
        sheet_main[f"F{start_row + i}"] = val[0]
        sheet_main[f"E{start_row + i}"] = val[1]
        sheet_main[f"E{start_row + i}"].number_format = number_format
        sheet_main[f"I{start_row + i}"] = val[2]
        sheet_main[f"G{start_row + i}"] = val[3]

    start_row += len(info['outcome'])
    for i, val in enumerate(info['income']):
        sheet_main[f"F{start_row + i}"] = val[0]
        sheet_main[f"D{start_row + i}"] = val[1]
        sheet_main[f"D{start_row + i}"].number_format = number_format
        sheet_main[f"G{start_row + i}"] = val[2]
        sheet_main[f'I{start_row + i}'] = info['tavr']   

    
def parse_tavr(sheet, income_start = 47, outcome_start = 35, income_len = 10, outcome_len = 10):
    """
        income_start - номер строчки в которой (Поступления в кассу) + 1
        income_len - Сколько столбцов доступно для зависи в (Поступления в кассу)

        outcome_start - номер строчки в которой (Затраты) + 1
        outcome_len - Сколько столбцов доступно для зависи в (Затраты)
    """

    income = []
    outcome = []
    count = 0
    while count < outcome_len:
        if sheet[f'B{outcome_start+count}'].value is not None:
            outcome.append([
                sheet[f'B{outcome_start+count}'].value, # наименование
                sheet[f'E{outcome_start+count}'].value, # сумма
                sheet[f'F{outcome_start+count}'].value, # тавр
                sheet[f'G{outcome_start+count}'].value # кому выданы деньги
            ])
        else:
            break
        count+=1

    count = 0
    while count < income_len:
        if sheet[f'B{income_start+count}'].value is not None:
            income.append([
                sheet[f'B{income_start+count}'].value, # наименование
                sheet[f'E{income_start+count}'].value, # сумма
                sheet[f'F{income_start+count}'].value # кому деньги
            ])
        else:
            break
        count+=1

    return income, outcome


def get_sheet(path_to_main, path_to_tavr, name):
    # Читает и возвращает листы нужные для обработки
    main_xlsx = openpyxl.load_workbook(path_to_main)
    sheet_main = main_xlsx[name]

    sheet_from_list = [openpyxl.load_workbook(path, data_only=True)['Лист1'] for path in path_to_tavr]

    # Сортируем по таврам
    sorted(sheet_from_list, key = lambda x: int(x['B1'].value[-1]))

    return main_xlsx, sheet_main, sheet_from_list


def get_aquaring(sheet_from, kkt = False):
    # Гиперпараметр для удобства смены начала чтения
    start_column = 22 + kkt

    aq = [ 
        sheet_from[f'D{start_column}'].value,
        sheet_from[f'D{start_column + 2}'].value,
        sheet_from[f'D{start_column + 4}'].value,
        sheet_from[f'D{start_column + 6}'].value,
        sheet_from[f'D{start_column + 8}'].value
    ]

    return aq


def get_revenue(sheet_from, kkt = False):
    # Гиперпараметр для удобства смены начала чтения
    start_column = 22 + kkt
    correction = 42 - kkt # почемуто ККТ идет до обычной штуки

    revenue = [
        [sheet_from[f'C{start_column}'].value, 'выручка'],
        [sheet_from[f'C{start_column + 2}'].value, 'выручка'],
        [sheet_from[f'C{start_column + 4}'].value, sheet_from[f'F{start_column + 4}'].value],
        [sheet_from[f'C{start_column + 6}'].value, sheet_from[f'F{start_column + 6}'].value],
        [sheet_from[f'C{start_column + 8}'].value, sheet_from[f'F{start_column + 8}'].value],
        [sheet_from[f'D{correction}'].value, 'выручка'],
    ]

    return revenue


def get_info(sheet_from):
    # Возвращает словарь с всей информацией
    info = dict()

    # ККТ будет отдельной матрицей
    info_kkt = dict()

    tavr = sheet_from['B1'].value
    tavr = 'т' + tavr[-1]

    # Тавр
    info['tavr'] = tavr
    info_kkt['tavr'] = tavr

    # Выручка
    info['revenue'] = get_revenue(sheet_from, kkt = False)
    info_kkt['revenue'] = get_revenue(sheet_from, kkt = True)

    # Эквайринг
    info['acquiring'] = get_aquaring(sheet_from, kkt = False)
    info_kkt['acquiring'] = get_aquaring(sheet_from, kkt = True)

    # Дата
    info['date'] = sheet_from['C2'].value
    info_kkt['date'] = sheet_from['C2'].value

    # Прибыль и траты
    info['income'], info['outcome'] = parse_tavr(sheet_from, income_start = 59, outcome_start = 46, income_len = 3, outcome_len = 5)
    info_kkt['income'], info_kkt['outcome'] = parse_tavr(sheet_from, income_start = 63, outcome_start = 52, income_len = 3, outcome_len = 5)


    # Наименования
    names = ["выручка ГСМ",
        "выручка АГЗС",
        "выручка магазин",
        "собственное пр-во",
        "выпечка",
        "округление при инкасации"]
    
    info['names'] = [name + " (н)" for name in names]
    info_kkt['names'] = [name + " ККТ" for name in names]

    return info, info_kkt


def new_cash(path_to_main, path_to_tavr, path_to_save, color):

    path_to_tavr_list = glob(os.path.join(path_to_tavr,'*.xlsx'))

    temp = len(path_to_tavr_list)
    if temp:
        logging.info(f"Всего {temp} отчета найдено по кассе")
    else:
        logging.error("Не найдено ни одного отчета по кассе, проверьте павильность пути, если отчеты были загружены.")
        return

    main_xlsx, sheet_main, sheet_from_list = get_sheet(path_to_main, path_to_tavr_list, name = 'нов касса')
    end = len(list(sheet_main.values)) + 1

    for i, sheet_from in enumerate(sheet_from_list):
        info, info_kkt = get_info(sheet_from)

        # Создаем шаблон
        start_row = end
        end = make_shablon(start_row, sheet_main, info)
        # Вставляем цифры
        insert_numbers(start_row, sheet_main, info)
        logging.info(f'Отчет номер {i+1} (н) добавлен')

        start_row = end
        end = make_shablon(start_row, sheet_main, info_kkt)
        # Вставляем цифры
        insert_numbers(start_row, sheet_main, info_kkt)
        logging.info(f'Отчет номер {i+1} ККТ добавлен')

    make_shablon(end, sheet_main, info, empty = True, color = color)
    logging.info(f'Шаблон добавлен')

    
    # Сохраняем обработанный файл main_xlsx
    main_xlsx.save(path_to_save)
    logging.info(f'Добавление и сохранение отчетов в новую кассу завершено')

